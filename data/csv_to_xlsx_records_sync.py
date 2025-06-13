import openpyxl
import csv
from openpyxl.utils import get_column_letter

CSV_PATH = 'data/records_and_roles.csv'
XLSX_PATH = 'data/records_and_roles.xlsx'
SHEET_NAME = 'records'

def read_csv_data(csv_path):
    with open(csv_path, encoding='utf-8') as f:
        reader = csv.reader(f)
        rows = list(reader)
    return rows

def update_xlsx_sheet_with_csv(xlsx_path, sheet_name, csv_rows):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[sheet_name]
    # 既存の書式を保ちつつ、ヘッダー以降のセル内容だけ書き換え
    for row_idx, csv_row in enumerate(csv_rows, 1):
        for col_idx, value in enumerate(csv_row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
    wb.save(xlsx_path)
    print(f"[INFO] {sheet_name}シートの内容をCSVで上書きしました: {xlsx_path}")

if __name__ == '__main__':
    csv_rows = read_csv_data(CSV_PATH)
    update_xlsx_sheet_with_csv(XLSX_PATH, SHEET_NAME, csv_rows)
