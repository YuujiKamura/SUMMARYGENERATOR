import json
import os
import pickle
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from pathlib import Path
import glob
import hashlib
import shutil
import tempfile
from PIL import Image
import logging
import traceback
from .thermometer_utils import assign_thermometer_remarks, thermometer_remarks_index, select_thermometer_remark, process_thermometer_remarks
from .caption_board_utils import judge_caption_board_closeup
from .record_matching_utils import is_thermometer_image, is_thermometer_or_caption_board_image
from .records_loader import load_records_from_json
from .path_manager import path_manager
from .chain_record_utils import ChainRecord, load_chain_records, find_chain_records_by_roles

# --- 設定 ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RECORDS_PATH = str(path_manager.default_records)  # 分類辞書
OUTPUT_PATH = os.path.abspath(os.path.join(BASE_DIR, '../output/summary_highres.xlsx'))
IMG_WIDTH = 1280  # CALモード画像サイズ
IMG_HEIGHT = 960
ROLE_MAPPING_PATH = str(path_manager.role_mapping)
CACHE_FILE_PREVIEW = os.path.abspath(os.path.join(BASE_DIR, 'detect_cache_preview.pkl'))

DATASET_JSON_PATH = str(path_manager.scan_for_images_dataset)

# --- ChainRecordリストのロード ---
chain_records = load_chain_records(RECORDS_PATH)
remarks_to_chain_record = {rec.remarks: rec for rec in chain_records if rec.remarks}

# --- マッピングロジック ---
logging.basicConfig(level=logging.DEBUG, filename="thermo_mapping_debug.log", filemode="w", encoding="utf-8")
logger = logging.getLogger(__name__)

def load_role_mapping():
    path = path_manager.role_mapping
    print(f"[role_mapping] ロードパス: {path}")
    if not path or not os.path.exists(path):
        print(f"[role_mapping] ファイルが存在しません: {path}")
        return {}
    try:
        with open(path, encoding='utf-8') as f:
            content = f.read()
            print(f"[role_mapping] ファイル内容:\n{content}")
            # BOMチェック
            if content and ord(content[0]) == 0xfeff:
                print("[role_mapping] BOM（Byte Order Mark）が検出されました")
            f.seek(0)
            mapping = json.load(open(path, encoding='utf-8'))
    except Exception as e:
        print(f"[role_mapping] JSONパースエラー: {e}")
        traceback.print_exc()
        return {}
    if not mapping:
        print(f"[role_mapping] ロード内容が空です: {path}")
    else:
        print(f"[role_mapping] ロード成功: {len(mapping)}件")
    return mapping

def load_image_roles_from_cache_file(cache_dir="src/image_preview_cache"):
    """
    image_preview_cache/*.json から画像パス→[role, ...] のdictを返す
    """
    result = {}
    if not os.path.exists(cache_dir):
        return result
    for fpath in glob.glob(os.path.join(cache_dir, "*.json")):
        try:
            with open(fpath, "r", encoding="utf-8") as f:
                data = json.load(f)
            img_path = data.get("image_path")
            bboxes = data.get("bboxes", [])
            roles = [b.get("role") for b in bboxes if b.get("role")]
            if img_path and roles:
                result[img_path] = roles
        except Exception as e:
            print(f"[WARN] {fpath} 読込失敗: {e}")
    return result

def get_photo_category_from_remarks(remarks: str) -> str:
    rec = remarks_to_chain_record.get(remarks)
    return rec.photo_category if rec else ''

def match_image_to_records(image_json_dict, records, mapping=None):
    """
    画像パス→キャッシュJSON（img_json）→ChainRecordリストのdictを返す（ワンストップマッチングフロー対応）
    各画像のキャッシュJSONデータ（img_json, 1画像分のdict）をそのまま判定に渡す。
    image_json_dict: {image_path: img_json, ...}
    """
    from summarygenerator.utils.record_matching_utils import match_roles_records_one_stop
    if mapping is None:
        from summarygenerator.utils.summary_generator import load_role_mapping
        mapping = load_role_mapping()
    result = {}
    for img_path, img_json in image_json_dict.items():
        matched = match_roles_records_one_stop(img_json, mapping, records)
        result[img_path] = matched
    return result

def export_highres_summary(image_json_dict, mapping, records, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "HighResSummary"
    ws.append(['画像', 'ワークカテゴリー', 'タイプ', 'サブタイプ', '備考', '割当ロール'])
    match_results = match_image_to_records(image_json_dict, records)
    for img_path, matched_records in match_results.items():
        img_json = image_json_dict.get(img_path, {})
        roles = img_json.get('roles', [])
        role_str = ", ".join(roles)
        if matched_records:
            for record in matched_records:
                ws.append(['', record.get('work_category',''), record.get('type',''), record.get('subtype',''), getattr(record, 'remarks', record.get('remarks', '')), role_str])
        else:
            ws.append(['', '', '', '', 'マッチなし', role_str])
        # 画像埋め込み
        if os.path.exists(img_path):
            try:
                img = XLImage(img_path)
                img.width, img.height = IMG_WIDTH, IMG_HEIGHT
                ws.add_image(img, f'A{ws.max_row}')
                ws.row_dimensions[ws.max_row].height = IMG_HEIGHT * 0.75
                ws.column_dimensions['A'].width = IMG_WIDTH / 7.5
            except Exception as e:
                ws.cell(row=ws.max_row, column=1, value=f'画像読込エラー: {e}')
        else:
            ws.cell(row=ws.max_row, column=1, value='画像ファイルなし')
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f'Excelサマリーを出力しました: {out_path}')

def get_all_image_data(json_path, folder_path, cache_dir, mapping=None):
    """
    画像リスト・各画像ごとのrolesリスト・remarksマッピング・親フォルダグループ・温度計ロールのローテーション割り当てまでを一括で返す
    個別画像JSON群（image_preview_cacheディレクトリ）を直接参照して構築する
    戻り値: dict
    {
        'image_json_dict': ...,  # 画像パス→img_json（1画像分のキャッシュJSON）
        'per_image_roles': ..., # 画像パス→[role, ...]（各画像ごとのrolesリスト）
        'match_results': ...,
        'folder_to_images': ...,
        'folder_names': ...,
        'thermo_remarks_map': ...,
        'records': ...,
    }
    """
    image_data = collect_image_data_from_cache(cache_dir)
    image_json_dict = image_data['image_json_dict']
    per_image_roles = image_data['per_image_roles']  # 各画像ごとのrolesリスト
    folder_to_images = image_data['folder_to_images']
    folder_names = image_data['folder_names']
    if mapping is None:
        mapping = load_role_mapping()
    # remarksマッピング
    match_results = match_image_to_records(image_json_dict, load_records_from_json(RECORDS_PATH))
    # thermo_remarks_map生成
    thermo_remarks_map = {}
    for img_path, remarks_list in match_results.items():
        abs_p = os.path.normcase(os.path.abspath(img_path))
        if remarks_list:
            thermo_remarks_map[abs_p] = remarks_list[0]
        else:
            thermo_remarks_map[abs_p] = None
    return {
        'image_json_dict': image_json_dict,
        'per_image_roles': per_image_roles,
        'match_results': match_results,
        'folder_to_images': folder_to_images,
        'folder_names': folder_names,
        'thermo_remarks_map': thermo_remarks_map,
        'records': load_records_from_json(RECORDS_PATH),
    }

def collect_image_data_from_cache(cache_dir):
    """
    image_preview_cacheディレクトリ内の全JSONを走査し、
    画像リスト・各画像ごとのrolesリスト・フォルダごとのリストを構築する。
    戻り値: dict
    {
        'image_json_dict': {image_path: img_json, ...},
        'per_image_roles': {image_path: [role, ...], ...}, # 各画像ごとのrolesリスト
        'folder_to_images': {folder_abs_path: [image_abs_path, ...], ...},
        'folder_names': [folder_abs_path, ...],
    }
    """
    import glob
    image_json_dict = {}
    per_image_roles = {}
    folder_to_images = {}
    if not os.path.exists(cache_dir):
        return {
            'image_json_dict': {},
            'per_image_roles': {},
            'folder_to_images': {},
            'folder_names': [],
        }
    for fpath in glob.glob(os.path.join(cache_dir, "*.json")):
        try:
            with open(fpath, "r", encoding="utf-8") as f:
                data = json.load(f)
            img_path = data.get("image_path")
            bboxes = data.get("bboxes", [])
            roles = [b.get("role") for b in bboxes if b.get("role")]
            if img_path:
                abs_img_path = os.path.abspath(img_path)
                image_json_dict[abs_img_path] = data
                per_image_roles[abs_img_path] = roles
                parent = os.path.abspath(os.path.dirname(img_path))
                folder_to_images.setdefault(parent, []).append(abs_img_path)
        except Exception as e:
            print(f"[WARN] {fpath} 読込失敗: {e}")
    for k in folder_to_images:
        folder_to_images[k].sort(key=lambda x: os.path.basename(x))
    folder_names = sorted(folder_to_images.keys())
    return {
        'image_json_dict': image_json_dict,
        'per_image_roles': per_image_roles,
        'folder_to_images': folder_to_images,
        'folder_names': folder_names,
    }

def collect_original_images_from_folder(folder_path):
    """
    指定フォルダ以下の全画像ファイル（jpg/png等）を再帰的に集め、
    folder_to_images, folder_names を返す
    """
    import os
    IMAGE_EXTS = ('.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tiff', '.webp')
    folder_to_images = {}
    folder_names = set()
    for root, _, files in os.walk(folder_path):
        images = [os.path.abspath(os.path.join(root, f)) for f in files if f.lower().endswith(IMAGE_EXTS)]
        if images:
            folder_to_images[os.path.abspath(root)] = images
            folder_names.add(os.path.abspath(root))
    return folder_to_images, sorted(folder_names)

# --- テスト用 ---
def main():
    mapping = load_role_mapping()
    # 画像パス→キャッシュJSON全体のdictを作る
    image_json_dict = {}
    cache_dir = "src/image_preview_cache"
    for fpath in glob.glob(os.path.join(cache_dir, "*.json")):
        try:
            with open(fpath, "r", encoding="utf-8") as f:
                data = json.load(f)
            img_path = data.get("image_path")
            if img_path:
                image_json_dict[img_path] = data
        except Exception as e:
            print(f"[WARN] {fpath} 読込失敗: {e}")
    records = load_records_from_json(RECORDS_PATH)
    export_highres_summary(image_json_dict, mapping, records, OUTPUT_PATH)

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--test-datadeploy", action="store_true")
    args = parser.parse_args()
    if args.test_datadeploy:
        from summarygenerator.utils.datadeploy_test import run_datadeploy_test
        from os.path import dirname, abspath, join
        BASE_DIR = dirname(abspath(__file__))
        DATASET_JSON_PATH = join(BASE_DIR, "scan_for_images_dataset.json")
        CACHE_DIR = join(BASE_DIR, "image_preview_cache")
        run_datadeploy_test(DATASET_JSON_PATH, CACHE_DIR, use_thermo_special=True)
    else:
        main()