import os
from pathlib import Path
import io
import math
import json
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Side, Font
from PIL import Image as PILImage
from .image_cache_utils import get_image_cache_path
from .debug_utils import generate_matching_debug_log

def export_excel_photobook(
    match_results, image_roles, records, out_path,
    parent_folder_name=None, cache_dir=None, role_mapping_path=None
):
    # role_mapping_pathが指定されていればmappingをロード
    if role_mapping_path:
        from .summary_generator import load_role_mapping
        load_role_mapping(role_mapping_path)
    wb = Workbook()
    n_rows = 9
    area_widths = [180, 15, 22, 22, 22]
    labels = ['番号', '工種', '種別', '細別', '測点', '備考']
    thin = Side(border_style="thin", color="000000")
    # 写真区分の優先順
    photo_category_order = ['施工状況写真', '出来形管理写真', '品質管理写真']
    # recordsのremarks→photo_category, remarks→datetime辞書を作成
    remarks_to_photo_category = {getattr(r, 'remarks', None): getattr(r, 'photo_category', '') for r in records if getattr(r, 'remarks', None)}
    remarks_to_datetime = {getattr(r, 'remarks', None): getattr(r, 'datetime', '') for r in records if getattr(r, 'remarks', None)}
    # 並び順：photo_category_order, datetime, 親ディレクトリ名, ファイル名
    def sort_key(item):
        img_path, matched_remarks = item
        remarks = matched_remarks[0] if matched_remarks else ''
        photo_category = remarks_to_photo_category.get(remarks, '')
        dt = remarks_to_datetime.get(remarks, '')
        parent_dir = Path(img_path).parent.name
        # 写真区分の優先順位
        try:
            cat_idx = photo_category_order.index(photo_category)
        except ValueError:
            cat_idx = 99
        return (cat_idx, photo_category, dt, parent_dir, img_path)
    sorted_items = sorted(
        match_results.items(),
        key=sort_key
    )
    img_total = len(sorted_items)
    page_count = math.ceil(img_total / 3)
    fixed_img_size = None
    # recordsからremarks→recordのdictを作成
    remarks_to_record = {getattr(r, 'remarks', None): r for r in records if getattr(r, 'remarks', None)}
    # ページ分割ロジック：写真区分が変わるたびに新しいページを作る
    # まず、sorted_itemsを「photo_categoryごとに3件ずつ」でページ分割
    pages = []
    current_category = None
    current_page = []
    for item in sorted_items:
        img_path, matched_remarks = item
        remarks = matched_remarks[0] if matched_remarks else ''
        photo_category = remarks_to_photo_category.get(remarks, '')
        if current_category is None:
            current_category = photo_category
        if photo_category != current_category or len(current_page) == 3:
            if current_page:
                pages.append((current_category, current_page))
            current_page = []
            current_category = photo_category
        current_page.append(item)
    if current_page:
        pages.append((current_category, current_page))
    # ページごとの処理
    photo_category_page_counter = {}
    for page_idx, (photo_category, page_items) in enumerate(pages):
        # ページカウンタ
        if photo_category not in photo_category_page_counter:
            photo_category_page_counter[photo_category] = 1
        else:
            photo_category_page_counter[photo_category] += 1
        page_num = photo_category_page_counter[photo_category]
        # シート名生成
        if photo_category:
            sheet_name = f"{photo_category.replace('写真','')}{page_num}"
        else:
            sheet_name = f"写真帳{page_idx+1}"
        if page_idx == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(sheet_name)
        # 1列目に幅15のスペーサー列を挿入
        ws.insert_cols(1)
        ws.column_dimensions['A'].width = 15
        # 1ページに収める印刷設定
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.horizontalCentered = True
        # 列幅設定（B列以降にarea_widthsを適用）
        for i, w in enumerate(area_widths):
            col_letter = chr(ord('A') + i + 1)  # B列=2列目から
            ws.column_dimensions[col_letter].width = w
        # 行高設定
        for i in range(1, 31):
            ws.row_dimensions[i].height = 81
        # 画像3枚分を出力
        for idx_in_page, item in enumerate(page_items):
            img_path, matched_remarks = item
            # --- 詳細ログ出力 ---
            class EntryLike:
                def __init__(self, path, location):
                    self.path = path
                    self.location = location

            # 測点（location）をキャッシュJSONから取得
            location = ''
            try:
                json_path = get_image_cache_path(
                    img_path, cache_dir
                ) if cache_dir else get_image_cache_path(img_path)
                if os.path.exists(json_path):
                    with open(json_path, 'r', encoding='utf-8') as f:
                        cache_data = json.load(f)
                    location = cache_data.get('location', '')
            except Exception:
                location = ''
            entry_like = EntryLike(img_path, location)
            debug_lines = generate_matching_debug_log(
                matched_remarks, entry_like
            )
            for line in debug_lines:
                print(line)
            row_starts = [2, 12, 22]
            row = row_starts[idx_in_page]
            # 画像貼り付け
            try:
                with PILImage.open(img_path) as pil_img:
                    pil_img = pil_img.convert('RGB')
                    orig_w, orig_h = pil_img.size
                    if fixed_img_size is None:
                        ratio = 1.0  # 100%
                        new_w = int(orig_w * ratio)
                        new_h = int(orig_h * ratio)
                        fixed_img_size = (new_w, new_h)
                    else:
                        new_w, new_h = fixed_img_size
                    pil_img = pil_img.resize((new_w, new_h))
                    img_bytes = io.BytesIO()
                    pil_img.save(img_bytes, format='JPEG', quality=70)
                    img_bytes.seek(0)
                    xl_img = XLImage(img_bytes)
                    xl_img.width, xl_img.height = new_w, new_h
                    ws.add_image(xl_img, f'B{row}')
                    ws.merge_cells(
                        start_row=row, start_column=2,
                        end_row=row+n_rows-1, end_column=2
                    )
                    ws.cell(row=row, column=2).alignment = Alignment(
                        horizontal='center', vertical='center'
                    )
            except Exception as e:
                ws.cell(row=row, column=2, value=f'画像読込エラー: {e}')
            else:
                ws.cell(row=row, column=2, value='画像ファイルなし')
            # キャプション出力 (ChainRecord 前提で簡素化)
            # matched_remarks は List[ChainRecord] を想定
            record = matched_remarks[0] if matched_remarks else None
            remarks = getattr(record, 'remarks', '') if record else ''

            # ChainRecord が無い場合に備えてダミー辞書
            if record is None:
                record_dict = {}
            else:
                # ChainRecord は to_dict() を実装しているはず
                record_dict = record.to_dict() if hasattr(record, 'to_dict') else dict(record.__dict__)
            subtype = record_dict.get('subtype', '')
            values = [
                str(idx_in_page + 1),
                record_dict.get('work_category', ''),
                record_dict.get('type', ''),
                record_dict.get('subtype', ''),
                location,
                remarks
            ]
            font28 = Font(size=28)
            for i in range(len(labels)):
                ws.cell(row=row+i, column=3, value=labels[i])
                ws.cell(row=row+i, column=4, value=str(values[i]))
                ws.cell(row=row+i, column=3).alignment = Alignment(
                    horizontal='left', vertical='center'
                )
                ws.cell(row=row+i, column=4).alignment = Alignment(
                    horizontal='left', vertical='center'
                )
                # すべてのセルのフォントサイズを28に
                for col in range(2, 7):
                    ws.cell(row=row+i, column=col).font = font28
            # B～E列の1～5行目相当に下罫線
            for i in range(6):
                for col in range(3, 7):
                    ws.cell(row=row+i, column=col).border = Border(bottom=thin)
            # 画像エントリの最後の行（9行目）のA～E列に下罫線
            for col in range(2, 7):
                ws.cell(row=row+8, column=col).border = Border(bottom=thin)
        # 余白設定を狭く
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.2
        ws.page_margins.bottom = 0.2
        ws.page_margins.header = 0.0
        ws.page_margins.footer = 0.0
        # 印刷範囲と水平中央配置を最後に設定
        ws.print_area = 'A1:F30'
        ws.page_setup.horizontalCentered = True
        # ページのB1セルに写真区分を記載し、ログも出力
        ws.cell(row=1, column=2, value=photo_category)
        ws.cell(row=1, column=2).font = Font(size=28)
        print(f"[PAGE] page={page_idx+1} photo_category={photo_category} sheet={sheet_name}")
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    if parent_folder_name:
        print(f'Excelテスト出力完了({parent_folder_name}): {out_path}')
    else:
        print(f'Excelテスト出力完了: {out_path}')
    print(f"[LOG] export_excel_photobook: records={len(records)}, match_results={len(match_results)}, out_path={out_path}")

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--test-datadeploy", action="store_true")
    args = parser.parse_args()
    if args.test_datadeploy:
        from .datadeploy_test import run_datadeploy_test
        from os.path import dirname, abspath, join
        BASE_DIR = dirname(abspath(__file__))
        DATASET_JSON_PATH = join(BASE_DIR, "scan_for_images_dataset.json")
        CACHE_DIR = join(BASE_DIR, "image_preview_cache")
        run_datadeploy_test(DATASET_JSON_PATH, CACHE_DIR, use_thermo_special=True)