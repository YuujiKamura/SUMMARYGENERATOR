import os
import sys
import pytest
from pathlib import Path
from PyQt6.QtWidgets import QApplication
from src.summary_generator_widget import SummaryGeneratorWidget
import tempfile
import shutil
from openpyxl import load_workbook
import PIL.Image

@pytest.fixture(scope="module")
def app():
    app = QApplication.instance() or QApplication([])
    yield app

@pytest.fixture
def temp_output_dir():
    d = tempfile.mkdtemp()
    yield d
    # shutil.rmtree(d)  # 一時的に削除を無効化

@pytest.mark.timeout(10)
def test_summary_excel_export_with_realdata(qtbot, app, temp_output_dir):
    """
    SummaryGeneratorWidgetで実データを使い、Excel出力が正しく行われるかE2Eテスト
    """
    # テストモードでウィジェット生成
    widget = SummaryGeneratorWidget(test_mode=True)
    qtbot.addWidget(widget)
    widget.show()
    qtbot.waitExposed(widget)

    # パスマネージャー経由で画像リストJSONを明示的に指定（コピーを使う）
    from src.utils.path_manager import PathManager
    path_manager = PathManager(write_protect=True)
    orig_json_path = str(path_manager.scan_for_images_dataset)
    assert os.path.exists(orig_json_path), f"画像リストJSONが存在しない: {orig_json_path}"
    temp_json = tempfile.mktemp(suffix='.json')
    shutil.copyfile(orig_json_path, temp_json)
    # グローバルなcurrent_image_list_jsonを保存
    orig_current_json = path_manager.current_image_list_json
    try:
        path_manager.current_image_list_json = temp_json
        widget.json_path_edit.setText(temp_json)
        widget.load_image_list_from_path_manager()
        qtbot.wait(1000)

        # Excel出力先
        out_path = os.path.join(temp_output_dir, "test_summary_export.xlsx")
        print(f"[E2E] Excel出力先: {out_path}")
        # 完了シグナルで待つ
        with qtbot.waitSignal(widget.test_finished, timeout=20000):
            widget.export_summary_to_path(out_path)
        assert os.path.exists(out_path), f"Excelファイルが出力されていない: {out_path}"

        # --- ここから温度管理写真のサイクル割り当てE2E検証 ---
        wb = load_workbook(out_path)
        ws = wb.active
        # ヘッダー行を取得
        headers = [cell.value for cell in ws[1]]
        remarks_idx = headers.index('備考')
        photo_cat_idx = headers.index('写真区分')
        # 品質管理写真カテゴリのremarksを抽出
        thermo_remarks = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[photo_cat_idx] == '品質管理写真':
                thermo_remarks.append(row[remarks_idx])
        # サイクル順序（到着→敷均し→初期締固→開放温度...）を検証
        from src.thermometer_utils import THERMO_REMARKS
        n = len(thermo_remarks)
        # 末尾3枚は開放温度
        assert n >= 4, "温度管理写真が4枚以上必要です"
        for i, rem in enumerate(thermo_remarks):
            if n - i <= 3:
                assert '開放温度' in rem, f"末尾3枚は開放温度であるべき: {rem}"
            else:
                expected = THERMO_REMARKS[i % 3]
                assert rem == expected, f"{i+1}枚目: {rem} != {expected}"
    finally:
        # テスト用JSONとグローバル状態を復元
        path_manager.current_image_list_json = orig_current_json
        os.remove(temp_json)